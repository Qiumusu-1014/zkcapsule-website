from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "assets" / "files" / "quotes"
OUT.mkdir(parents=True, exist_ok=True)

BLUE = "126BFF"
BLUE_DARK = "0D4ED8"
NAVY = "0B1220"
PAPER = "F4F8FF"
PAPER_ALT = "FAFCFF"
LINE = "D7E2F1"
MUTED = "53657E"
WHITE = "FFFFFF"
CHECKED = "☑"
UNCHECKED = "☐"

THIN_LINE = Side(style="thin", color=LINE)
BORDER = Border(left=THIN_LINE, right=THIN_LINE, top=THIN_LINE, bottom=THIN_LINE)


COMMON_OPTIONS = [
    ("全套卫浴 / 带洗浴", "Complete bathroom / shower", 25000),
    ("储物柜（衣物收纳）", "Storage cabinet / wardrobe", 7000),
    ("床及床垫（单套）", "Bed and mattress set", 9300),
    ("第二床位 / 第二床垫", "Second bed / second mattress", 3200),
    ("桌椅套装", "Desk and chair set", 2500),
    ("电器包", "Appliance package", 40000),
    ("电动窗帘", "Motorized curtains", 5700),
    ("Low-E 玻璃升级", "Low-E glass upgrade", 20000),
    ("天窗", "Skylight", 15000),
    ("内饰升级包", "Interior finish upgrade package", 40000),
    ("地暖系统", "Floor heating system", 20000),
    ("灯带 / 氛围照明系统", "Light strip / ambient lighting system", 2500),
    ("中央空调", "Central air conditioning", 40000),
    ("挂机空调", "Wall-mounted air conditioner", 4700),
    ("新风系统", "Fresh air ventilation system", 15000),
    ("顶部太阳能光伏板", "Roof solar PV panels", 25000),
    ("储能电池系统", "Energy storage battery system", 50000),
    ("逆变器 / 能源管理系统", "Inverter / energy management system", 20000),
    ("星链连接硬件及安装辅材", "Starlink hardware and installation accessories", 33200),
    ("国际通信系统", "International communication system", 10800),
    ("LED 大屏幕", "LED display screen", 28000),
    ("投影仪", "Projector", 5800),
    ("3D 展示系统", "3D presentation system", 25000),
    ("AI 智能全系统", "AI smart system", 100000),
    ("AI 语音控制 / 设备联动全屋智能", "AI voice control / device-linked smart system", 20000),
    ("净水系统", "Water purification system", 10000),
    ("智能门锁", "Smart door lock", 1600),
    ("监控 / 烟感 / 水浸报警基础包", "Basic security package", 4700),
    ("运输成本合计", "Estimated logistics cost", 96700),
]

ZK05_SPECIAL_OPTIONS = [
    ("露台户外休闲包", "Terrace outdoor leisure package", 85000, "户外露台、休闲座椅、遮阳及景观布置", "Outdoor terrace, leisure seating, shading and landscape setup"),
    ("活动运营移动部署包", "Mobile deployment package for events", 120000, "牵引、支撑腿、户外接口及快速转场结构", "Towing, support legs, outdoor interfaces and quick relocation structure"),
    ("户外遮阳棚 / 雨棚", "Outdoor canopy / awning", 28000, "适合露台、营地和景区休闲区", "Suitable for terraces, campsites and resort leisure areas"),
    ("户外家具与景观灯", "Outdoor furniture and landscape lighting", 22000, "露台桌椅、氛围灯和景观照明", "Terrace furniture, ambient lighting and landscape lights"),
]

ZK06_SPECIAL_OPTIONS = [
    ("外置楼梯与连廊包", "External staircase and corridor package", 85000, "双舱固定叠加结构已包含，本项仅为外部通行升级", "The fixed stacked dual-capsule structure is included; this is only an access upgrade"),
    ("观景露台与安全栏杆包", "Viewing terrace and safety railing package", 70000, "二层观景露台、栏杆和户外休闲区", "Second-level viewing terrace, railings and outdoor leisure area"),
    ("酒店套房软装包", "Hotel suite soft furnishing package", 120000, "大床、迷你吧、衣柜、软装和度假客房氛围", "Large bed, minibar, wardrobe, soft furnishing and resort-suite atmosphere"),
    ("竖向管线检修增强包", "Vertical MEP maintenance enhancement", 35000, "上下层水电暖通检修和集中接口增强", "Enhanced vertical MEP access and centralized service interfaces"),
]

ZK07_SPECIAL_OPTIONS = [
    ("商业窗口与折叠柜台包", "Service window and folding counter package", 35000, "外摆售卖窗口、折叠柜台和营业开口", "Outdoor-facing service window, folding counter and commercial opening"),
    ("发光招牌 / 品牌门头", "Illuminated signage / branded fascia", 25000, "景区、商业街或展会品牌识别", "Brand identity for scenic areas, retail streets or exhibitions"),
    ("POS 收银与网络包", "POS and network package", 18000, "POS 位、收银网络和基础布线", "POS location, cashier network and basic wiring"),
    ("水吧操作台 / 展示柜包", "Beverage counter / display cabinet package", 45000, "水吧操作台、展示柜、冰箱及储物", "Beverage counter, display cabinet, refrigerator and storage"),
]

ZK08_SPECIAL_OPTIONS = [
    ("卫星通信增强包", "Satellite communication enhancement package", 120000, "在基础应急通信上增加卫星/多链路冗余", "Adds satellite / multi-link redundancy to the base emergency communication system"),
    ("大型地图指挥屏增强包", "Large map command display enhancement", 80000, "地图大屏、指挥显示和多源画面接入", "Large map display, command screen and multi-source video access"),
    ("便携应急医疗包", "Portable emergency medical package", 60000, "基础救援、急救和应急医疗物资配置", "Basic rescue, first-aid and emergency medical supplies"),
    ("无人机指挥 / 充电接口", "Drone command / charging interface", 90000, "无人机起降、充电和应急巡检接口", "Drone landing, charging and emergency inspection interfaces"),
    ("额外储能 / 发电机接口", "Extra storage / generator interface", 150000, "高规格备用电源和外接发电机接口增强", "Enhanced backup power and external generator interface"),
]

ZK09_SPECIAL_OPTIONS = [
    ("多舱组合深化设计包", "Multi-capsule combination design package", 180000, "双舱/多舱组合平面、结构和动线深化", "Detailed planning, structure and circulation for dual/multi-capsule combinations"),
    ("旗舰外立面定制包", "Flagship facade customization package", 250000, "高端定制外立面、品牌形象和夜景灯光", "Premium customized facade, brand identity and night lighting"),
    ("VIP 会客 / 吧台升级包", "VIP lounge / bar counter upgrade package", 180000, "独立会客区、吧台和商务接待软装", "Private lounge area, bar counter and business reception furnishing"),
    ("顶级内饰材料包", "Premium interior materials package", 300000, "高级饰面、定制家具和酒店级材料", "Premium finishes, custom furniture and hotel-grade materials"),
    ("全屋智能 / 灯光联动包", "Whole-space smart control / lighting package", 160000, "全屋智能、灯光场景和设备联动", "Whole-space smart control, lighting scenes and device linkage"),
    ("品牌展厅多媒体包", "Brand showroom multimedia package", 220000, "品牌展示、LED/投影/互动多媒体系统", "Brand display, LED / projection and interactive multimedia system"),
]

MODELS = [
    {
        "code": "ZK01",
        "zh_name": "基础生活舱",
        "en_name": "Basic Living Capsule",
        "zh_desc": "参考 28㎡",
        "en_desc": "Approx. 28 sqm reference configuration",
        "base": 450000,
        "special_options": [],
    },
    {
        "code": "ZK02",
        "zh_name": "绿色度假舱",
        "en_name": "Green Resort Capsule",
        "zh_desc": "绿色豪华版配置",
        "en_desc": "Green luxury configuration",
        "base": 570000,
        "special_options": [],
    },
    {
        "code": "ZK03",
        "zh_name": "文化教育舱",
        "en_name": "Culture & Education Capsule",
        "zh_desc": "文化宣传 / 教育培训配置",
        "en_desc": "Culture / education configuration",
        "base": 700000,
        "special_options": [],
    },
    {
        "code": "ZK04",
        "zh_name": "医疗服务舱",
        "en_name": "Medical Service Capsule",
        "zh_desc": "诊疗医药配置",
        "en_desc": "Medical service configuration",
        "base": 850000,
        "special_options": [],
    },
    {
        "code": "ZK05",
        "zh_name": "露台舱",
        "en_name": "Terrace Capsule",
        "zh_desc": "露台、户外休闲、景区度假基础配置",
        "en_desc": "Terrace, outdoor leisure and resort-use base configuration",
        "base": 620000,
        "special_options": ZK05_SPECIAL_OPTIONS,
    },
    {
        "code": "ZK06",
        "zh_name": "叠加度假套房舱",
        "en_name": "Stacked Resort Suite Capsule",
        "zh_desc": "固定双舱叠加结构，基础价按 2 × ZK01 基础生活舱计算",
        "en_desc": "Fixed two-capsule stacked structure; base price equals 2 × ZK01 Basic Living Capsule",
        "base": 900000,
        "special_options": ZK06_SPECIAL_OPTIONS,
    },
    {
        "code": "ZK07",
        "zh_name": "商业服务舱",
        "en_name": "Commercial Service Capsule",
        "zh_desc": "轻量商业运营舱，适合售卖窗口、快闪商业和服务点",
        "en_desc": "Light commercial operation capsule for service windows, pop-up retail and service points",
        "base": 200000,
        "special_options": ZK07_SPECIAL_OPTIONS,
    },
    {
        "code": "ZK08",
        "zh_name": "应急指挥舱",
        "en_name": "Emergency Command Capsule",
        "zh_desc": "高规格应急系统基础配置，含应急通信、备用电源、指挥调度和警示识别能力",
        "en_desc": "High-spec emergency system base configuration with emergency communications, backup power, command dispatch and warning identity",
        "base": 1500000,
        "special_options": ZK08_SPECIAL_OPTIONS,
    },
    {
        "code": "ZK09",
        "zh_name": "旗舰定制舱",
        "en_name": "Flagship Customized Capsule",
        "zh_desc": "双舱/多舱组合、VIP会客、吧台、顶级内饰和全屋智能基础配置",
        "en_desc": "Dual/multi-capsule combination with VIP lounge, bar counter, premium interior and smart control base configuration",
        "base": 1680000,
        "special_options": ZK09_SPECIAL_OPTIONS,
    },
    {"code": "ABADI", "zh_name": "高端定制智慧太空舱", "en_name": "Premium Customized Smart Capsule", "zh_desc": "项目定制", "en_desc": "Customized project", "base": None, "special_options": []},
]

MOBILE_OPTION = ("移动款附加项", "Mobile version add-on", 120000)


def localized(model, lang):
    if lang == "zh":
        return model["zh_name"], model["zh_desc"], "中文报价单"
    return model["en_name"], model["en_desc"], "English Quotation"


def labels(lang):
    if lang == "zh":
        return {
            "company": "公司",
            "contact": "联系人",
            "address": "地址",
            "currency": "币种",
            "base": "基础价格",
            "options": "选配合计",
            "total": "报价合计",
            "base_section": "基础项目",
            "item": "项目",
            "price": "价格",
            "select": "勾选",
            "amount": "计入金额",
            "remark": "备注",
            "optional_section": "选配项目",
            "contract": "合同价为准",
            "rmb": "人民币 RMB",
            "empty_base": "",
        }
    return {
        "company": "Company",
        "contact": "Contact",
        "address": "Address",
        "currency": "Currency",
        "base": "Base Price",
        "options": "Options Total",
        "total": "Quotation Total",
        "base_section": "Base Item",
        "item": "Item",
        "price": "Price",
        "select": "Select",
        "amount": "Included Amount",
        "remark": "Remark",
        "optional_section": "Optional Items",
        "contract": "Contract price prevails",
        "rmb": "RMB",
        "empty_base": "",
    }


def option_parts(option, lang):
    if len(option) == 3:
        zh, en, amount = option
        return (zh if lang == "zh" else en), amount, ""
    if len(option) == 5:
        zh, en, amount, zh_remark, en_remark = option
        return (zh if lang == "zh" else en), amount, zh_remark if lang == "zh" else en_remark
    raise ValueError(f"Unexpected option format: {option}")


def set_font(cell, size=10, bold=False, color=NAVY):
    cell.font = Font(name="Microsoft YaHei", size=size, bold=bold, color=color)


def set_currency(cell):
    cell.number_format = '¥#,##0'
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)


def write_summary(ws, model, lang, label):
    ws["A5"] = label["base"]
    ws["B5"] = label["options"]
    ws["C5"] = label["total"]
    ws["D5"] = label["contract"]
    ws["E5"] = ""

    ws["A6"] = model["base"] if model["base"] is not None else None
    ws["B6"] = ""
    ws["C6"] = '=IF(SUM(A6:B6)=0,"",SUM(A6:B6))'
    ws["D6"] = label["rmb"]
    ws["E6"] = ""

    for col in range(1, 6):
        head = ws.cell(5, col)
        val = ws.cell(6, col)
        head.fill = PatternFill("solid", fgColor=NAVY)
        val.fill = PatternFill("solid", fgColor=PAPER)
        set_font(head, 10, True, WHITE)
        set_font(val, 12, True, BLUE_DARK)
        head.alignment = Alignment(horizontal="center", vertical="center")
        val.alignment = Alignment(horizontal="center", vertical="center")
        if col in (1, 2, 3):
            set_currency(val)


def add_data_validation(ws, start_row, end_row):
    dv = DataValidation(type="list", formula1=f'"{UNCHECKED},{CHECKED}"', allow_blank=False)
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid value"
    dv.prompt = "Select checkbox"
    dv.promptTitle = "Checkbox"
    ws.add_data_validation(dv)
    dv.add(f"C{start_row}:C{end_row}")


def apply_sheet_style(ws, option_start, option_end):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{option_start}"

    widths = {"A": 42, "B": 18, "C": 12, "D": 18, "E": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            set_font(cell)

    for row_index in range(1, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 26

    ws.row_dimensions[1].height = 38
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 34
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[option_start - 1].height = 30

    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for row_index in (8, option_start - 1):
        for col in range(1, 6):
            cell = ws.cell(row_index, col)
            cell.fill = PatternFill("solid", fgColor=BLUE)
            set_font(cell, 10, True, WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(2, 5):
        for col in range(1, 6):
            ws.cell(row_index, col).fill = PatternFill("solid", fgColor=PAPER_ALT)
        set_font(ws.cell(row_index, 1), 10, True, MUTED)
        set_font(ws.cell(row_index, 3), 10, True, MUTED)

    for row_index in range(9, ws.max_row + 1):
        if row_index % 2 == 0:
            for col in range(1, 6):
                ws.cell(row_index, col).fill = PatternFill("solid", fgColor=PAPER_ALT)

    for row_index in range(option_start, option_end + 1):
        ws.cell(row_index, 3).font = Font(name="Segoe UI Symbol", size=13, bold=True, color=BLUE_DARK)
        ws.cell(row_index, 3).alignment = Alignment(horizontal="center", vertical="center")
        set_currency(ws.cell(row_index, 2))
        set_currency(ws.cell(row_index, 4))

    for row_index in (9,):
        set_currency(ws.cell(row_index, 2))

    ws.conditional_formatting.add(
        f"A{option_start}:E{option_end}",
        FormulaRule(formula=[f'$C{option_start}="{CHECKED}"'], fill=PatternFill("solid", fgColor="EAF2FF")),
    )
    ws.auto_filter.ref = f"A{option_start - 1}:E{option_end}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def append_sheet(ws, model, lang):
    name, desc, title_suffix = localized(model, lang)
    label = labels(lang)
    code = model["code"]

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{code} {name} {title_suffix}"

    ws.append([""])
    ws.append([label["company"], "中科国际 Zhongke International", label["contact"], "杨涛峰 / 13319274381", ""])
    ws.append([label["address"], "香港尖沙咀麼地道75号南洋中心2座7楼", label["currency"], label["rmb"], ""])

    write_summary(ws, model, lang, label)

    ws.append([""])
    ws.append([label["base_section"], label["price"], "", "", label["remark"]])
    ws.append([f"{code} {name}", model["base"] if model["base"] is not None else None, "", "", desc])
    ws.append(["", "", "", "", ""])
    ws.append([label["optional_section"], label["price"], label["select"], label["amount"], label["remark"]])

    option_start = ws.max_row + 1
    options = list(model.get("special_options", [])) + list(COMMON_OPTIONS)
    if code in {"ZK01", "ZK02", "ZK03", "ZK04"}:
        options.insert(0, MOBILE_OPTION)

    for option in options:
        item, amount, remark = option_parts(option, lang)
        row = ws.max_row + 1
        ws.append([item, amount, UNCHECKED, f'=IF($C{row}="{CHECKED}",$B{row},"")', remark])

    option_end = ws.max_row
    ws["B6"] = f'=IF(SUM(D{option_start}:D{option_end})=0,"",SUM(D{option_start}:D{option_end}))'
    ws["C6"] = '=IF(SUM(A6:B6)=0,"",SUM(A6:B6))'

    add_data_validation(ws, option_start, option_end)
    apply_sheet_style(ws, option_start, option_end)


def build_quote(model):
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws_zh = wb.active
    ws_zh.title = "中文报价单"
    append_sheet(ws_zh, model, "zh")

    ws_en = wb.create_sheet("English Quotation")
    append_sheet(ws_en, model, "en")

    path = OUT / f"{model['code']}-quotation.xlsx"
    wb.save(path)
    return path


def verify_quote(path):
    wb = load_workbook(path, data_only=False)
    if wb.sheetnames != ["中文报价单", "English Quotation"]:
        raise ValueError(f"{path.name}: unexpected sheets {wb.sheetnames}")

    banned_terms = ("报价说明", "Quotation Note", "按项目实际配置确认")
    for ws in wb.worksheets:
        flat_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
        joined_values = "\n".join(flat_values)
        for term in banned_terms:
            if term in joined_values:
                raise ValueError(f"{path.name} {ws.title}: banned text found: {term}")

        if ws.max_column != 5:
            raise ValueError(f"{path.name} {ws.title}: expected 5 columns, got {ws.max_column}")

        option_header_row = None
        for row_index in range(1, ws.max_row + 1):
            if ws.cell(row_index, 3).value in {"勾选", "Select"}:
                option_header_row = row_index
                break
        if option_header_row is None:
            raise ValueError(f"{path.name} {ws.title}: option header missing")

        option_start = option_header_row + 1
        option_end = ws.max_row
        expected_total = f'=IF(SUM(D{option_start}:D{option_end})=0,"",SUM(D{option_start}:D{option_end}))'
        if ws["B6"].value != expected_total:
            raise ValueError(f"{path.name} {ws.title}: wrong option total formula")
        if ws["C6"].value != '=IF(SUM(A6:B6)=0,"",SUM(A6:B6))':
            raise ValueError(f"{path.name} {ws.title}: wrong final total formula")
        if ws.cell(option_start, 3).value != UNCHECKED:
            raise ValueError(f"{path.name} {ws.title}: unchecked default missing")
        expected_row_formula = f'=IF($C{option_start}="{CHECKED}",$B{option_start},"")'
        if ws.cell(option_start, 4).value != expected_row_formula:
            raise ValueError(f"{path.name} {ws.title}: option row formula missing")

        validation_ranges = [str(item.sqref) for item in ws.data_validations.dataValidation]
        expected_range = f"C{option_start}:C{option_end}"
        if not any(expected_range in item for item in validation_ranges):
            raise ValueError(f"{path.name} {ws.title}: checkbox validation missing")


if __name__ == "__main__":
    built_files = []
    for item in MODELS:
        path = build_quote(item)
        verify_quote(path)
        built_files.append(path)
        print(path)
    print(f"verified {len(built_files)} quotation workbooks")
